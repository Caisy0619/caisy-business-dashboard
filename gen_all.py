"""生成3家公司的36个月经营数据 (2023-01 ~ 2025-12)"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import random
import math

random.seed(42)

COMPANIES = {
    '鼎盛医疗': {  # 持续增长型
        'revenue_base': 180, 'revenue_trend': 0.018,  # 月均增长1.8%
        'cost_ratio': 0.70, 'selling_ratio': 0.10, 'admin_ratio': 0.08, 'finance_ratio': 0.006,
        'salary_base': 48, 'salary_growth': 0.008,
        'headcount_base': 42, 'headcount_growth': 0.005,
    },
    '恒达器械': {  # 平稳型
        'revenue_base': 165, 'revenue_trend': 0.002,
        'cost_ratio': 0.72, 'selling_ratio': 0.11, 'admin_ratio': 0.09, 'finance_ratio': 0.007,
        'salary_base': 45, 'salary_growth': 0.003,
        'headcount_base': 40, 'headcount_growth': 0.002,
    },
    '锐新科技': {  # 下滑型
        'revenue_base': 200, 'revenue_trend': -0.012,
        'cost_ratio': 0.75, 'selling_ratio': 0.13, 'admin_ratio': 0.11, 'finance_ratio': 0.008,
        'salary_base': 55, 'salary_growth': 0.001,
        'headcount_base': 50, 'headcount_growth': -0.003,
    },
}

MONTHS = []
for y in range(2023, 2026):
    for m in range(1, 13):
        MONTHS.append(f"{y}-{m:02d}")

def make_data(cfg):
    """生成36个月的数据（万元为单位，返回列表）"""
    revs, costs, sells, admins, finances, profits, salaries, hcs = [], [], [], [], [], [], [], []
    
    for i in range(36):
        # 营收带季节性波动
        season = 1 + 0.08 * math.sin(i * math.pi / 6)  # 6个月周期
        noise = 1 + random.gauss(0, 0.03)
        rev = cfg['revenue_base'] * (1 + cfg['revenue_trend']) ** i * season * noise
        rev = round(rev * 10000)  # 转为元
        
        cost = round(rev * cfg['cost_ratio'] * (1 + random.gauss(0, 0.02)))
        sell = round(rev * cfg['selling_ratio'] * (1 + random.gauss(0, 0.05)))
        admin = round(rev * cfg['admin_ratio'] * (1 + random.gauss(0, 0.05)))
        finance = round(rev * cfg['finance_ratio'] * (1 + random.gauss(0, 0.1)))
        
        # 其他数据
        sal = round(cfg['salary_base'] * 10000 * (1 + cfg['salary_growth']) ** i * (1 + random.gauss(0, 0.02)))
        hc = max(20, round(cfg['headcount_base'] * (1 + cfg['headcount_growth']) ** i))
        
        # 净利润 = 营收 - 成本 - 费用 + 其他收益(忽略)
        pft = rev - cost - sell - admin - finance
        
        revs.append(rev)
        costs.append(cost)
        sells.append(sell)
        admins.append(admin)
        finances.append(finance)
        profits.append(pft)
        salaries.append(sal)
        hcs.append(hc)
    
    return revs, costs, sells, admins, finances, profits, salaries, hcs

# 准备输出目录
import os
base = "/Users/caisy/Desktop/经营看板"
os.makedirs(base, exist_ok=True)

thin = Side(style='thin', color='C0C8B8')
border = Border(bottom=thin)
header_fill = PatternFill(start_color="E8F0E0", end_color="E8F0E0", fill_type="solid")

for name, cfg in COMPANIES.items():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "经营数据"
    
    # 公司名称（第2行B列）
    ws.cell(row=1, column=1, value="公司名称").font = Font(bold=True, size=10, color="6B7F6B")
    ws.cell(row=1, column=2, value=name).font = Font(bold=True, size=12, color="3A5C3A")
    
    # 空行
    ws.cell(row=2, column=1, value="")
    
    # 表头：科目 | 2023-01 | 2023-02 | ...
    row_idx = 3
    ws.cell(row=row_idx, column=1, value="科目").font = Font(bold=True, size=10)
    ws.cell(row=row_idx, column=1).fill = header_fill
    ws.cell(row=row_idx, column=1).border = border
    for j, m in enumerate(MONTHS, 2):
        cell = ws.cell(row=row_idx, column=j, value=m)
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal='center', text_rotation=0)
        cell.fill = header_fill
        cell.border = border
        cell.number_format = '@'  # 文本格式
    
    # 数据
    data = make_data(cfg)
    labels = ['营业收入', '营业成本', '销售费用', '管理费用', '财务费用', '净利润', '薪酬总额', '员工人数']
    
    for i, (label, values) in enumerate(zip(labels, data)):
        r = row_idx + 1 + i
        cell = ws.cell(row=r, column=1, value=label)
        cell.font = Font(bold=False, size=10)
        cell.border = border
        for j, val in enumerate(values, 2):
            cell = ws.cell(row=r, column=j, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal='right')
            if label == '员工人数':
                cell.number_format = '0'
            else:
                cell.number_format = '#,##0'
    
    # 列宽
    ws.column_dimensions['A'].width = 14
    for j in range(2, 38):
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = 11
    
    safe_name = name
    path = os.path.join(base, f"{safe_name}.xlsx")
    wb.save(path)
    print(f"✅ {name}: 36月数据已保存 → {path}")
