import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "经营数据"

# 简单两列：项目 + 金额
headers = ["项目", "金额"]
for i, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=i, value=h)
    cell.font = Font(bold=True, size=11)
    cell.alignment = Alignment(horizontal='center')
    cell.fill = PatternFill(start_color="E8F0E0", end_color="E8F0E0", fill_type="solid")

data = [
    ("营业收入", 22000000),
    ("营业成本", 15200000),
    ("销售费用", 2200000),
    ("管理费用", 1740000),
    ("财务费用", 145000),
    ("净利润", 2012250),
    ("薪酬总额", 6250000),
    ("员工人数", 45),
]

thin = Side(style='thin', color='D0D8C8')
border = Border(bottom=thin)

for i, (name, val) in enumerate(data, 2):
    ws.cell(row=i, column=1, value=name).font = Font(bold=False, size=11)
    ws.cell(row=i, column=2, value=val).number_format = '#,##0'
    ws.cell(row=i, column=2).alignment = Alignment(horizontal='right')
    for j in [1, 2]:
        ws.cell(row=i, column=j).border = border

ws.column_dimensions['A'].width = 16
ws.column_dimensions['B'].width = 16

note = ws.cell(row=len(data)+3, column=1, value="模拟数据：年营收2200万，员工45人，薪酬625万")
note.font = Font(color="888888", size=10)

wb.save("/Users/caisy/Desktop/经营看板/样例数据.xlsx")
print("Done")
